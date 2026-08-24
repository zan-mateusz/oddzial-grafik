"""Import grafiku z arkusza .xlsx/.xlsm — z automatycznym rozpoznaniem układu."""
from __future__ import annotations

import datetime as dt
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

MAX_SCAN_ROWS = 60
MAX_SCAN_COLS = 60


@dataclass
class SheetGrid:
    """Surowa zawartość arkusza jako tekst."""

    name: str
    cells: list[list[str]]

    def value(self, row: int, col: int) -> str:
        if 0 <= row < len(self.cells) and 0 <= col < len(self.cells[row]):
            return self.cells[row][col]
        return ""

    @property
    def n_rows(self) -> int:
        return len(self.cells)

    @property
    def n_cols(self) -> int:
        return max((len(r) for r in self.cells), default=0)


@dataclass
class Layout:
    """Gdzie w arkuszu znajdują się nazwiska i dni miesiąca."""

    header_row: int = -1
    first_data_row: int = -1
    name_col: int = -1
    day_cols: dict[int, int] = field(default_factory=dict)  # dzień -> indeks kolumny
    confidence: float = 0.0

    @property
    def ok(self) -> bool:
        return self.header_row >= 0 and self.name_col >= 0 and len(self.day_cols) >= 20


SPREADSHEET_SUFFIXES = (".xlsx", ".xlsm", ".ods")


def read_sheets(path: str | Path) -> list[SheetGrid]:
    """Wczytuje arkusz Excela (.xlsx) albo OpenDocument (.ods)."""
    if Path(path).suffix.lower() == ".ods":
        from app.io.ods_import import read_ods_sheets

        return read_ods_sheets(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    grids = []
    for ws in wb.worksheets:
        cells: list[list[str]] = []
        for r, row in enumerate(ws.iter_rows(max_row=MAX_SCAN_ROWS, max_col=MAX_SCAN_COLS)):
            cells.append([_as_text(c.value) for c in row])
        grids.append(SheetGrid(ws.title, cells))
    wb.close()
    return grids


def best_sheet_index(grids: list[SheetGrid], expected_days: int = 31) -> int:
    """Wskazuje arkusz, który naprawdę zawiera grafik.

    Skoroszyty prowadzone latami mają zwykle kilka pustych arkuszy ("Arkusz14")
    przed właściwym. Bez tego wyboru kreator otwierałby pierwszy z brzegu, pusty.
    """
    best_index, best_score = 0, (-1.0, -1)
    for i, grid in enumerate(grids):
        layout = detect_layout(grid, expected_days)
        rows = len(extract_rows(grid, layout)) if layout.ok else 0
        score = (layout.confidence if layout.ok else 0.0, rows)
        if score > best_score:
            best_index, best_score = i, score
    return best_index


def _as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _day_number(text: str) -> int | None:
    """Wyciąga numer dnia z nagłówka: "1", "1 Pn", "01.09", datę."""
    text = text.strip()
    if not text:
        return None
    m = re.match(r"^(\d{1,2})\s*$", text)
    if m:
        n = int(m.group(1))
        return n if 1 <= n <= 31 else None
    m = re.match(r"^(\d{1,2})[\s.\-/]", text)
    if m:
        n = int(m.group(1))
        return n if 1 <= n <= 31 else None
    m = re.match(r"^\d{4}-(\d{2})-(\d{2})$", text)
    if m:
        return int(m.group(2))
    return None


def detect_layout(grid: SheetGrid, expected_days: int = 31) -> Layout:
    """expected_days: liczba dni importowanego miesiąca (do oceny trafności)."""
    """Szuka wiersza z kolejnymi numerami dni oraz kolumny z nazwiskami."""
    best = Layout()
    for r in range(min(grid.n_rows, MAX_SCAN_ROWS)):
        day_cols: dict[int, int] = {}
        for c in range(grid.n_cols):
            day = _day_number(grid.value(r, c))
            if day is not None and day not in day_cols:
                day_cols[day] = c
        # Nagłówek dni to taki, który zawiera 1..N w rosnących kolumnach.
        run = _longest_increasing_run(day_cols)
        if len(run) < 20:
            continue
        confidence = len(run) / expected_days
        if confidence > best.confidence:
            name_col = _detect_name_column(grid, r, min(run.values()))
            best = Layout(
                header_row=r,
                first_data_row=_detect_first_data_row(grid, r, name_col),
                name_col=name_col,
                day_cols=run,
                confidence=confidence,
            )
    return best


def _longest_increasing_run(day_cols: dict[int, int]) -> dict[int, int]:
    """Zostawia dni ułożone rosnąco w rosnących kolumnach.

    Luki są dopuszczalne — przy odczycie ze zdjęcia pojedynczy numer bywa
    nieczytelny, a przy arkuszach zdarzają się scalone komórki. Wymóg
    rosnących kolumn nadal odrzuca przypadkowe liczby rozsiane po arkuszu.
    """
    out: dict[int, int] = {}
    prev_col = -1
    for day in sorted(day_cols):
        col = day_cols[day]
        if col > prev_col:
            out[day] = col
            prev_col = col
    return _fill_missing_days(out)


def _fill_missing_days(day_cols: dict[int, int]) -> dict[int, int]:
    """Uzupełnia brakujące dni, gdy kolumny są równo rozstawione.

    Dzięki temu jeden numer zgubiony przez OCR nie oznacza utraty całej kolumny
    z dyżurami.
    """
    if len(day_cols) < 4:
        return day_cols
    days = sorted(day_cols)
    steps = [
        (day_cols[b] - day_cols[a]) / (b - a)
        for a, b in zip(days, days[1:]) if b - a <= 3
    ]
    if not steps:
        return day_cols
    step = sorted(steps)[len(steps) // 2]
    if step <= 0:
        return day_cols
    filled = dict(day_cols)
    for day in range(days[0], days[-1] + 1):
        if day in filled:
            continue
        prev = max((d for d in days if d < day), default=None)
        if prev is None:
            continue
        filled[day] = round(day_cols[prev] + step * (day - prev))
    return dict(sorted(filled.items()))


def _detect_name_column(grid: SheetGrid, header_row: int, first_day_col: int) -> int:
    """Kolumna z największą liczbą tekstów wyglądających na nazwiska."""
    best_col, best_score = 0, -1
    for c in range(min(first_day_col, grid.n_cols)):
        score = 0
        for r in range(header_row + 1, min(grid.n_rows, header_row + 40)):
            text = grid.value(r, c)
            if len(text) >= 3 and any(ch.isalpha() for ch in text) and not text.isdigit():
                score += 1
        if score > best_score:
            best_col, best_score = c, score
    return best_col


def _detect_first_data_row(grid: SheetGrid, header_row: int, name_col: int) -> int:
    for r in range(header_row + 1, grid.n_rows):
        if grid.value(r, name_col).strip():
            return r
    return header_row + 1


_ORDINAL_PREFIX = re.compile(r"^\s*\d{1,3}\s*[.)]\s*")


def strip_ordinal(name: str) -> str:
    """Usuwa numer porządkowy z początku nazwiska ("1. Kowalska" -> "Kowalska")."""
    return _ORDINAL_PREFIX.sub("", name).strip()


def normalize_name(name: str) -> str:
    """Do porównywania nazwisk: bez ogonków, wielkości liter i inicjałów."""
    text = unicodedata.normalize("NFKD", strip_ordinal(name).lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z\s]", " ", text)
    parts = [p for p in text.split() if len(p) > 1]
    return " ".join(sorted(parts))


# Rdzenie polskich nazw miesięcy — pasują też do form odmienionych
# ("STYCZNIA", "MAJA"), bo dopasowanie działa na początku wyrazu.
_MONTH_STEMS = [
    ("stycz", 1), ("lut", 2), ("mar", 3), ("kwie", 4), ("maj", 5), ("czerw", 6),
    ("lip", 7), ("sierp", 8), ("wrze", 9), ("pazdzier", 10), ("listopad", 11),
    ("grud", 12),
]


def guess_month(*sources: str) -> tuple[int | None, int | None]:
    """Odgaduje (rok, miesiąc) z nazwy arkusza lub pliku, np. "CZERWIEC 2026".

    Nazwa arkusza jest pewniejsza niż nazwa pliku: przygotowując kolejny grafik
    użytkownik zwykle kopiuje plik, a nazwę arkusza zmienia na bieżący miesiąc.
    """
    month = year = None
    for text in sources:
        if not text:
            continue
        plain = unicodedata.normalize("NFKD", text.lower())
        plain = "".join(c for c in plain if not unicodedata.combining(c))
        if month is None:
            for stem, number in _MONTH_STEMS:
                if stem in plain:
                    month = number
                    break
        if year is None:
            found = re.search(r"(20\d{2})", plain)
            if found:
                year = int(found.group(1))
    return year, month


@dataclass
class ImportedRow:
    source_name: str
    entries: dict[int, str]           # dzień -> treść komórki
    employee_id: int | None = None    # dopasowany pracownik
    create_new: bool = False

    @property
    def filled(self) -> int:
        return sum(1 for v in self.entries.values() if v.strip())


def extract_rows(grid: SheetGrid, layout: Layout, max_rows: int = 60) -> list[ImportedRow]:
    rows: list[ImportedRow] = []
    blanks = 0
    for r in range(layout.first_data_row, min(grid.n_rows, layout.first_data_row + max_rows)):
        name = grid.value(r, layout.name_col).strip()
        if not name:
            blanks += 1
            if blanks >= 3:
                break
            continue
        blanks = 0
        name = strip_ordinal(name)
        if not name or _looks_like_summary(name) or _looks_like_shift_code(name):
            continue
        entries = {
            day: grid.value(r, col).strip() for day, col in layout.day_cols.items()
        }
        # Wiersz bez żadnego wpisu w kolumnach dni to stopka, notatka albo pusty
        # wiersz ozdobny — nie pracownik.
        if not any(v for v in entries.values()):
            continue
        rows.append(ImportedRow(source_name=name, entries=entries))
    return rows


_SUMMARY_WORDS = (
    "razem", "suma", "łącznie", "lacznie", "ogółem", "ogolem", "obsada",
    "legenda", "objaśnienia", "objasnienia", "podpis", "sporządził", "sporzadzil",
)


def _looks_like_summary(name: str) -> bool:
    low = name.lower()
    return any(word in low for word in _SUMMARY_WORDS)


def _looks_like_shift_code(name: str) -> bool:
    """Wiersze legendy zaczynają się od kodu zmiany (D, N, UŻ, L4) — nie od nazwiska.

    Nazwiska zawierają małe litery, kody są krótkie i pisane wersalikami.
    """
    stripped = name.strip()
    return len(stripped) <= 6 and not any(ch.islower() for ch in stripped)


def normalize_token(text: str) -> str:
    """Pojedynczy wyraz bez ogonków i wielkości liter — do porównań nazwisk."""
    plain = unicodedata.normalize("NFKD", text.strip().lower())
    plain = "".join(ch for ch in plain if not unicodedata.combining(ch))
    return re.sub(r"[^a-z]", "", plain)


def match_employees(rows: list[ImportedRow], employees: list) -> None:
    """Dopasowuje wiersze do pracowników po pełnym imieniu i nazwisku.

    Gdy pełne dopasowanie zawiedzie (w arkuszu bywa "Kowalska A."), próbujemy
    po samym nazwisku — ale wyłącznie wtedy, gdy wskazuje ono jednoznacznie na
    jedną osobę. Dopasowanie po imieniu byłoby niebezpieczne: dwie Ewy na
    oddziale to norma, a pomyłka przypisałaby dyżury nie tej osobie.
    """
    exact: dict[str, int] = {}
    by_surname: dict[str, set[int]] = {}
    for emp in employees:
        full = f"{emp['last_name']} {emp['first_name']}".strip()
        exact.setdefault(normalize_name(full), emp["id"])
        surname = normalize_token(emp["last_name"])
        if surname:
            by_surname.setdefault(surname, set()).add(emp["id"])

    for row in rows:
        key = normalize_name(row.source_name)
        emp_id = exact.get(key)
        if emp_id is None:
            tokens = set(key.split())
            candidates: set[int] = set()
            for surname, ids in by_surname.items():
                if surname in tokens:
                    candidates |= ids
            # Przy dwóch osobach o tym samym nazwisku wybór zostawiamy użytkownikowi.
            if len(candidates) == 1:
                emp_id = candidates.pop()
        row.employee_id = emp_id
        row.create_new = emp_id is None


def apply_import(
    db, year: int, month: int, rows: list[ImportedRow], replace: bool = True
) -> tuple[int, int]:
    """Zapisuje zaimportowane wiersze. Zwraca (liczba wpisów, nowi pracownicy)."""
    from app.core.calendar_pl import month_days

    valid_days = {d.day: d for d in month_days(year, month)}
    created = 0
    items: list[tuple[int, dt.date, str]] = []

    for row in rows:
        emp_id = row.employee_id
        if emp_id is None:
            if not row.create_new:
                continue
            last, first = _split_name(row.source_name)
            emp_id = db.add_employee(last, first)
            created += 1
        for day, text in row.entries.items():
            date = valid_days.get(day)
            if date is None:
                continue
            items.append((emp_id, date, text))

    if replace:
        db.clear_month(year, month)
    db.set_entries_bulk(items)
    return sum(1 for _, _, t in items if t.strip()), created


def _split_name(text: str) -> tuple[str, str]:
    parts = [p for p in re.split(r"[\s,]+", strip_ordinal(text)) if p]
    if not parts:
        return text.strip(), ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])
