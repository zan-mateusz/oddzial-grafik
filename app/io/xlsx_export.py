"""Eksport grafiku do pliku .xlsx (Excel / LibreOffice)."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.calendar_pl import (
    DayKind, PL_MONTHS_TITLE, PL_WEEKDAYS_SHORT, day_kind, holiday_name,
    month_days, month_norm,
)
from app.core.rules import load_rules
from app.core.shifts import (
    Category, Entry, ShiftType, fmt_days_hours, fmt_minutes, resolve,
)
from app.core.stats import summarize_month

THIN = Side(style="thin", color="B0B6C0")
MEDIUM = Side(style="medium", color="6B7280")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = {
    DayKind.WEEKDAY: "F3F4F6",
    DayKind.SATURDAY: "DCE7F5",
    DayKind.SUNDAY: "C9DCF1",
    DayKind.HOLIDAY: "F6CFD8",
}
HEADER_FONT_COLOR = {
    DayKind.WEEKDAY: "374151",
    DayKind.SATURDAY: "1E4B7A",
    DayKind.SUNDAY: "12385E",
    DayKind.HOLIDAY: "8C1D32",
}
DAY_FILL = {
    DayKind.WEEKDAY: None,
    DayKind.SATURDAY: "EDF3FA",
    DayKind.SUNDAY: "E4EDF8",
    DayKind.HOLIDAY: "FBE7EC",
}

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

# Kolumny podsumowania: (klucz, nagłówek, szerokość). Zestaw jest taki sam jak
# na ekranie — dobierany do zawartości miesiąca.
NARROW = 10
WIDE = 13


def _headers(multi: bool, has_sick: bool):
    columns = [("wymiar", "Wymiar", NARROW)]
    if multi:
        columns += [("glowne", "Dyż. gł.", WIDE), ("zastepcze", "Dyż. zast.", WIDE)]
    else:
        columns.append(("glowne", "Dyżury", WIDE))
    columns += [
        ("bilans", "Bilans", NARROW),
        ("dzien", "Dzień", WIDE),
        ("noc", "Noc", WIDE),
        ("swieta", "Święta", WIDE),
        ("urlop", "Urlop", WIDE),
    ]
    if has_sick:
        columns.append(("l4", "L4", WIDE))
    return columns


def _hex(color: str) -> str:
    return color.lstrip("#").upper()[:6] or "FFFFFF"


def export_month(
    path: str | Path,
    db,
    year: int,
    month: int,
    ward_name: str = "",
) -> Path:
    """Zapisuje grafik miesiąca do .xlsx — każde piętro na osobnej karcie."""
    path = Path(path)
    days = month_days(year, month)
    types = db.shift_types_by_code()
    rules = load_rules(db)
    norm = month_norm(year, month, rules.daily_norm_minutes)

    # Sumy miesięczne liczymy ze wszystkich pięter — pracownik ma jeden wymiar
    # czasu pracy niezależnie od tego, gdzie odbył dyżur.
    all_entries = _resolved(db.month_entries(year, month), types)
    entry_floors = db.month_entry_floors(year, month)
    has_sick = any(e.category is Category.SICK for e in all_entries.values())

    floors = db.floors() or [None]
    wb = Workbook()
    wb.remove(wb.active)

    for floor in floors:
        floor_id = floor["id"] if floor is not None else None
        floor_label = floor["name"] if floor is not None else ""
        employees = db.employees_for_month(year, month, floor_id)
        floor_entries = _resolved(db.month_entries(year, month, floor_id), types)
        # Podział na dyżury własne i zastępcze jest cechą pracownika, więc
        # liczymy go z wpisów całego miesiąca, nie tylko z tego piętra.
        home = {e["id"]: e["floor_id"] for e in employees}
        main_entries, cover_entries = {}, {}
        for key, entry in all_entries.items():
            where = entry_floors.get(key)
            target = main_entries if where == home.get(key[0]) else cover_entries
            target[key] = entry

        month_summaries = summarize_month(year, month, employees, all_entries, rules)
        main_summaries = summarize_month(year, month, employees, main_entries, rules)
        cover_summaries = summarize_month(year, month, employees, cover_entries, rules)

        ws = wb.create_sheet(_sheet_title(floor_label, year, month))
        _write_title(ws, year, month, norm, ward_name, len(days), floor_label)
        header_row = 4
        multi = len(floors) > 1 and floor is not None
        _write_headers(ws, header_row, days, multi, has_sick)
        last_row = _write_body(
            ws, header_row + 1, days, employees, floor_entries,
            month_summaries, main_summaries, cover_summaries, multi,
            floor_id, has_sick,
        )
        _write_legend(ws, last_row + 2, db.shift_types(), multi)
        _apply_layout(ws, days, header_row, last_row, multi, has_sick)

    wb.save(path)
    return path


def _resolved(raw: dict, types: dict) -> dict:
    out = {}
    for key, text in raw.items():
        entry = resolve(text, types)
        if entry is not None:
            out[key] = entry
    return out


def _sheet_title(floor_label: str, year: int, month: int) -> str:
    """Nazwa karty — Excel zabrania kilku znaków i ogranicza długość do 31."""
    base = floor_label or f"Grafik {month:02d}-{year}"
    for bad in "[]:*?/\\":
        base = base.replace(bad, "-")
    return base[:31]


def _write_title(ws: Worksheet, year: int, month: int, norm, ward: str, n_days: int,
                 floor_label: str = "") -> None:
    total_cols = 2 + n_days + 9
    heading = f"Grafik dyżurów — {PL_MONTHS_TITLE[month - 1]} {year}"
    if floor_label:
        heading += f" — {floor_label}"
    ws.cell(row=1, column=1, value=heading)
    ws.cell(row=1, column=1).font = Font(size=15, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 12))

    subtitle = (
        f"Wymiar czasu pracy: {fmt_minutes(norm.minutes)} h "
        f"({norm.working_days} dni roboczych)"
    )
    if ward:
        subtitle = f"{ward}    •    {subtitle}"
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).font = Font(size=10, color="4B5563")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(total_cols, 12))


def _write_headers(ws: Worksheet, row: int, days: list[dt.date], multi: bool = False,
                   has_sick: bool = False) -> None:
    ws.cell(row=row, column=1, value="Nazwisko i imię")
    ws.cell(row=row, column=2, value="Etat")
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="E5E7EB")
        c.alignment = CENTER_WRAP
        c.border = BORDER

    for i, day in enumerate(days):
        col = 3 + i
        kind = day_kind(day)
        c = ws.cell(row=row, column=col, value=f"{day.day}\n{PL_WEEKDAYS_SHORT[day.weekday()]}")
        c.font = Font(bold=True, size=9, color=HEADER_FONT_COLOR[kind])
        c.fill = PatternFill("solid", fgColor=HEADER_FILL[kind])
        c.alignment = CENTER_WRAP
        c.border = BORDER
        name = holiday_name(day)
        if name:
            c.comment = _comment(name)

    start = 3 + len(days)
    for i, (key, label, _) in enumerate(_headers(multi, has_sick)):
        c = ws.cell(row=row, column=start + i, value=label)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill(
            "solid", fgColor="DCE6F1" if key in ("wymiar", "bilans") else "E5E7EB"
        )
        c.alignment = CENTER_WRAP
        c.border = BORDER


def _comment(text: str):
    from openpyxl.comments import Comment
    c = Comment(text, "Grafik")
    c.width, c.height = 180, 40
    return c


def _summary_values(month, main, cover) -> dict[str, str]:
    """Te same liczby, które widać na ekranie."""
    if month is None:
        return {}
    return {
        "wymiar": month.norm_hhmm,
        "glowne": fmt_days_hours(main.shift_days, main.worked_minutes) if main else "",
        "zastepcze": fmt_days_hours(cover.shift_days, cover.worked_minutes)
                     if cover else "",
        "bilans": month.balance_hhmm,
        "dzien": fmt_days_hours(month.day_shifts, month.day_minutes),
        "noc": fmt_days_hours(month.night_shifts, month.night_shift_minutes),
        "swieta": fmt_days_hours(month.holidays_worked, month.holiday_minutes),
        "urlop": fmt_days_hours(month.leave_days, month.leave_minutes),
        "l4": fmt_days_hours(month.sick_days, month.sick_minutes),
    }


def _write_body(ws, first_row, days, employees, entries, summaries,
                main_summaries=None, cover_summaries=None, multi=False,
                floor_id=None, has_sick=False) -> int:
    row = first_row
    for emp in employees:
        name = f"{emp['last_name']} {emp['first_name']}".strip()
        cover = multi and emp["floor_id"] != floor_id
        if cover:
            name += "  (zastępstwo)"
        c = ws.cell(row=row, column=1, value=name)
        c.alignment = LEFT
        c.font = Font(size=10, italic=cover, color="8C4A00" if cover else "000000")
        c.border = BORDER

        fte = "1/1" if emp["fte_num"] == emp["fte_den"] else f"{emp['fte_num']}/{emp['fte_den']}"
        c = ws.cell(row=row, column=2, value=fte)
        c.alignment = CENTER
        c.font = Font(size=9, color="6B7280")
        c.border = BORDER

        for i, day in enumerate(days):
            col = 3 + i
            entry: Entry | None = entries.get((emp["id"], day))
            cell = ws.cell(row=row, column=col, value=entry.label if entry else None)
            cell.alignment = CENTER
            cell.border = BORDER
            if entry is not None:
                cell.font = Font(size=9, bold=entry.category is Category.WORK,
                                 color="B00020" if entry.unknown else "000000")
                cell.fill = PatternFill("solid", fgColor=_hex(entry.color))
                if entry.minutes:
                    cell.comment = _comment(f"{fmt_minutes(entry.minutes)} h")
            else:
                fill = DAY_FILL[day_kind(day)]
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

        values = _summary_values(
            summaries.get(emp["id"]),
            (main_summaries or {}).get(emp["id"]),
            (cover_summaries or {}).get(emp["id"]),
        )
        month = summaries.get(emp["id"])
        start = 3 + len(days)
        for i, (key, _, _) in enumerate(_headers(multi, has_sick)):
            cell = ws.cell(row=row, column=start + i, value=values.get(key) or None)
            cell.alignment = CENTER
            cell.border = BORDER
            summary_col = key in ("wymiar", "bilans")
            cell.fill = PatternFill(
                "solid", fgColor="EFF3F8" if summary_col else "F7F8FA"
            )
            colour = "000000"
            if key == "bilans" and month:
                colour = "B45309" if month.balance_minutes > 0 else (
                    "B00020" if month.balance_minutes < 0 else "15803D")
            elif key == "zastepcze":
                colour = "8C4A00"
            cell.font = Font(size=9, bold=key == "bilans", color=colour)
        row += 1
    return row - 1


def _write_legend(ws: Worksheet, row: int, shift_types: list[ShiftType],
                  multi: bool = False) -> None:
    ws.cell(row=row, column=1, value="Legenda").font = Font(bold=True, size=10)
    row += 1
    col = 1
    for st in shift_types:
        code_cell = ws.cell(row=row, column=col, value=st.code)
        code_cell.fill = PatternFill("solid", fgColor=_hex(st.color))
        code_cell.alignment = CENTER
        code_cell.border = BORDER
        code_cell.font = Font(size=9, bold=True)

        desc = st.name or ""
        if st.start and st.end:
            desc += f" ({st.start.strftime('%H:%M')}–{st.end.strftime('%H:%M')})"
        label = ws.cell(row=row, column=col + 1, value=desc.strip())
        label.font = Font(size=9)
        label.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 4)

        col += 6
        if col > 25:
            col = 1
            row += 1
    row += 2
    text = ("Kolory nagłówków: sobota i niedziela — odcienie niebieskiego, "
            "święto ustawowe — różowy. Na oddziale dyżury pełnione są również w te dni.")
    if multi:
        text += (
            " Osoby oznaczone jako zastępstwo pracują na innym piętrze."
            " Kolumny Godziny, Wymiar i Bilans obejmują cały ich miesiąc,"
            " a Godz. tu i Dyż. tu — tylko to piętro."
        )
    note = ws.cell(row=row, column=1, value=text)
    note.font = Font(size=9, italic=True, color="4B5563")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)


def _apply_layout(ws: Worksheet, days, header_row: int, last_row: int,
                  multi: bool = False, has_sick: bool = False) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 6
    for i in range(len(days)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 4.6
    start = 3 + len(days)
    for i, (_, _, width) in enumerate(_headers(multi, has_sick)):
        ws.column_dimensions[get_column_letter(start + i)].width = width

    ws.row_dimensions[header_row].height = 30
    for r in range(header_row + 1, last_row + 1):
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_options.horizontalCentered = True
